import csv
import threading
from threading import Thread
import numpy as np
import argparse
import cv2
import xlwt
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter import ttk  # Normal Tkinter.* widgets are not themed!
from ttkthemes import ThemedTk
from PIL import Image, ImageTk
from openpyxl import Workbook
import tkinter.font as font
import datetime
from tkinter.messagebox import showinfo
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl.styles import PatternFill

from Services.Critere import Critere
from Services.Historique_comptage import Historique_comptage
from Services.HistoriqueComtpageService import HistoriqueComtpageService
from Services.User import User
from Services.UserService import UserService
from Services.Video import Video
from Vue.MyDatePicker import MyDatePicker
from counting.Point import Point
from counting.centroidtracker import CentroidTracker
from counting.trackableobject import TrackableObject
from tkcalendar import *
# Initialize the parameters
confThreshold = 0.6  # Confidence threshold
nmsThreshold = 0.4  # Non-maximum suppression threshold
inpWidth = 416  # Width of network's input image
inpHeight = 416  # Height of network's input image

# initialize the frame dimensions (we'll set them as soon as we read
# the first frame from the video)
W = None
H = None

# instantiate our centroid tracker, then initialize a list to store
# each of our dlib correlation trackers, followed by a dictionary to
# map each unique object ID to a TrackableObject
ct = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct2 = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct3 = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct4 = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct5 = CentroidTracker(maxDisappeared=40, maxDistance=50)
trackers = []
trackableObjects = {}
trackableObjects1 = {}
trackableObjects2 = {}
trackableObjects3 = {}
trackableObjects4 = {}
trackableObjects5 = {}
# initialize the total number of frames processed thus far, along
# with the total number of objects that have moved either up or down
totalDown = 0
totalUp = 0
totalDownCar = 0
totalUpCar = 0
totalDownBus = 0
totalUpBus = 0
totalDownTruck = 0
totalUpTruck = 0
totalDownMotor = 0
totalUpMotor = 0

showCar=True
showPerson=True
showTruck=True
showBus=True
showMotorBike=True

isEntree=False
isSortie=False

canDrawEnter=0
canDrawOut=0

# cap = cv2.VideoCapture("video.mp4")
# construct the argument parse and parse the arguments
ap = argparse.ArgumentParser()

ap.add_argument('-l', '--labels', type=str, default='model/coco.names', help='Path to label file')
# ap.add_argument('--video', help='Path to video file.', default='D:/Stage/67 HA/D2/DJI_0004.MP4')

args = ap.parse_args()

# initialize the list of class labels MobileNet SSD was trained to
# detect, then generate a set of bounding box colors for each class
CLASSES = ["background", "aeroplane", "bicycle", "bird", "boat",
           "bottle", "bus", "car", "cat", "chair", "cow", "diningtable",
           "dog", "horse", "motorbike", "person", "pottedplant", "sheep",
           "sofa", "train", "tvmonitor"]
classesFile = open(args.labels).read().strip().split('\n')

# Give the configuration and weight files for the model and load the network using them.
modelConfiguration = "model/yolov3.cfg";
modelWeights = "model/yolov3.weights";

# load our serialized model from disk
net = cv2.dnn.readNetFromDarknet(modelConfiguration, modelWeights)
net.setPreferableBackend(cv2.dnn.DNN_BACKEND_OPENCV)
net.setPreferableTarget(cv2.dnn.DNN_TARGET_OPENCL)
COLORS = np.random.uniform(0, 255, size=(len(CLASSES), 3))
video = ""
cap = ""
fps = 0
frame_count = 0
duration = 0
userProfile=""
videoList = []
canvasList=[]
histoDataList=[]
videoAttenteList=[]
def getOutputsNames(net):
    # Get the names of all the layers in the network
    layersNames = net.getLayerNames()
    # Get the names of the output layers, i.e. the layers with unconnected outputs
    return [layersNames[i[0] - 1] for i in net.getUnconnectedOutLayers()]


def draw_bounding_boxes(image, boxes, confidences, classIDs, idxs, colors):
    if len(idxs) > 0:
        for i in idxs.flatten():
            # extract bounding box coordinates
            x, y = boxes[i][0], boxes[i][1]
            w, h = boxes[i][2], boxes[i][3]

            # draw the bounding box and label on the image
            color = [int(c) for c in colors[classIDs[i]]]
            cv2.rectangle(image, (x, y), (x + w, y + h), color, 2)

            text = "{}: {:.4f}".format(classesFile[classIDs[i]], confidences[i])
            cv2.putText(image, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

    return image


# Draw the predicted bounding box
def drawPred(classId, conf, left, top, right, bottom, frame):
    # Draw a bounding box.
    global pointSortie1
    global pointSortie2
    if(classId==0):
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
    if (classId == 2):
        cv2.rectangle(frame, (left, top), (right, bottom), (255, 0, 0), 2)
    if (classId == 6):
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
    if (classId == 5):
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 255), 2)
    if (classId == 3):
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 165, 255), 2)
    # Draw a center of a bounding box
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]

    cv2.line(frame, (pointSortie1.x, pointSortie1.y), (pointSortie2.x, pointSortie2.y), (0, 47, 255), 2)
    #cv2.line(frame, (0, frameHeight // 2 - 50), (frameWidth, frameHeight // 2 - 50), (0, 255, 255), 2)
    cv2.circle(frame, (left + (right - left) // 2, top + (bottom - top) // 2), 3, (0, 0, 255), -1)

    counter = []
    coun = 0
    if (top + (bottom - top) // 2 in range(frameHeight // 2 - 2, frameHeight // 2 + 2)):
        coun += 1
        # print(coun)

        counter.append(coun)
    # label = 'Pedestrians: '.format(str(counter))
    text = "{}".format(classesFile[classId])

   # cv2.putText(frame, text, (left, top - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 1)


def getOnlyOnSegmentPoints( segment_first_extremity,  segment_last_extremity,  points,  maximal_allowed_distance):

    segment_first_point_x = segment_first_extremity.x
    segment_first_point_y = segment_first_extremity.y
    segment_last_point_x = segment_last_extremity.x
    segment_last_point_y = segment_last_extremity.y
    test_x = points.x
    test_y = points.y
    returned = True
    k_numerator = (test_x - segment_first_point_x) * (segment_last_point_x - segment_first_point_x)
    + (test_y - segment_first_point_y) * (segment_last_point_y - segment_first_point_y)

    k_denominator = (segment_last_point_x - segment_first_point_x) * (segment_last_point_x - segment_first_point_x)
    + (segment_last_point_y - segment_first_point_y) * (segment_last_point_y - segment_first_point_y);
    k = k_numerator / k_denominator;


    p =  Point((k * (segment_last_point_x - segment_first_point_x) + (segment_first_point_x)), (k * (segment_last_point_y - segment_first_point_y) + (segment_first_point_y)));

    if (k < 0.0 and Point.distanceBetweenTwoPoint(points, segment_first_extremity) > maximal_allowed_distance):
        return False


    if (k >= 0.0  and k <= 1.0 and Point.distanceBetweenTwoPoint(points, p) > maximal_allowed_distance):
        return False


    if (k > 1.0 and Point.distanceBetweenTwoPoint(points, segment_last_extremity) > maximal_allowed_distance):
        return False

    return returned;

def postprocess2(frame, outs):
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]
    rects = []

    # Scan through all the bounding boxes output from the network and keep only the
    # ones with high confidence scores. Assign the box's class label as the class with the highest score.
    classIds = []
    confidences = []
    boxes = []
    for out in outs:
        for detection in out:
            scores = detection[5:]
            classId = np.argmax(scores)
            confidence = scores[classId]
            if confidence > confThreshold:
                center_x = int(detection[0] * frameWidth)
                center_y = int(detection[1] * frameHeight)
                width = int(detection[2] * frameWidth)
                height = int(detection[3] * frameHeight)
                left = int(center_x - width / 2)
                top = int(center_y - height / 2)
                classIds.append(classId)
                confidences.append(float(confidence))
                boxes.append([left, top, width, height])

    # Perform non maximum suppression to eliminate redundant overlapping boxes with
    # lower confidences.
    indices = cv2.dnn.NMSBoxes(boxes, confidences, confThreshold, nmsThreshold)
    for i in indices:
        i = i[0]
        box = boxes[i]
        left = box[0]
        top = box[1]
        width = box[2]
        height = box[3]

        if classIds[i] == 2 and showCar:
            rects.append((left, top, left + width, top + height))
            objects2 = ct2.update(rects)
            countingObject2(objects2, trackableObjects1, "car", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
            continue
        if classIds[i] == 3 and showMotorBike:
            rects.append((left, top, left + width, top + height))
            objects3 = ct3.update(rects)
            countingObject2(objects3, trackableObjects1, "motorbike", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
            continue
        if classIds[i] == 5 and showBus:
            rects.append((left, top, left + width, top + height))
            # use the centroid tracker to associate the (1) old object
            # centroids with (2) the newly computed object centroids
            objects4 = ct4.update(rects)
            countingObject2(objects4, trackableObjects1, "bus", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
            continue
        if classIds[i] == 6 and showTruck:
            rects.append((left, top, left + width, top + height))
            # use the centroid tracker to associate the (1) old object
            # centroids with (2) the newly computed object centroids
            objects5 = ct5.update(rects)
            countingObject2(objects5, trackableObjects1, "truck", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
            continue

def countingObject(objects, trackableObjectList, type, frame):
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]

    global totalDown
    global totalUp
    global totalUpCar
    global totalUpBus
    global totalUpMotor
    global totalUpTruck
    global totalDownCar
    global totalDownBus
    global totalDownTruck
    global totalDownMotor
    # loop over the tracked objects
    for (objectID, centroid) in objects.items():
        # check to see if a trackable object exists for the current
        # object ID
        to = trackableObjectList.get(objectID, None)

        # if there is no existing trackable object, create one
        if to is None:
            to = TrackableObject(objectID, centroid)

        # otherwise, there is a trackable object so we can utilize it
        # to determine direction
        else:
            # the difference between the y-coordinate of the *current*
            # centroid and the mean of *previous* centroids will tell
            # us in which direction the object is moving (negative for
            # 'up' and positive for 'down')
            y = [c[1] for c in to.centroids]
            direction = centroid[1] - np.mean(y)
            to.centroids.append(centroid)

            # check to see if the object has been counted or not
            if not to.counted:
                # if the direction is negative (indicating the object
                # is moving up) AND the centroid is above the center
                # line, count the object

                if direction < 0 and centroid[1] in range(frameHeight // 2 - 30, frameHeight // 2 + 30):
                    if type is "person":
                        totalUp += 1
                    if type is "car":
                        totalUpCar += 1
                    if type is "bus":
                        totalUpBus += 1
                    if type is "motorbike":
                        totalDownMotor += 1
                    if type is "truck":
                        totalUpTruck += 1
                    to.counted = True

                # if the direction is positive (indicating the object
                # is moving down) AND the centroid is below the
                # center line, count the object
                elif direction > 0 and centroid[1] in range(frameHeight // 2 - 30, frameHeight // 2 + 30):
                    if type is "person":
                        totalDown += 1
                    if type is "car":
                        totalDownCar += 1
                    if type is "bus":
                        totalDownBus += 1
                    if type is "motorbike":
                        totalDownMotor += 1
                    if type is "truck":
                        totalDownTruck += 1
                    to.counted = True

        # store the trackable object in our dictionary
        trackableObjectList[objectID] = to
        # draw both the ID of the object and the centroid of the
        # object on the output frame
        # text = "ID {}".format(objectID)
        # cv.putText(frame, text, (centroid[0] - 10, centroid[1] - 10),
        # cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
        cv2.circle(frame, (centroid[0], centroid[1]), 4, (0, 255, 0), -1)
    # construct a tuple of information we will be displaying on the
    # frame
    #info = [
        #     ("Person :", str(totalUp) + "/" + str(totalDown)),
        #    ("Car  :", str(totalUpCar) + "/" + str(totalDownCar)),
        #    ("Bus  :", str(totalUpBus) + "/" + str(totalDownBus)),
        #    ("Motorbike  :", str(totalUpMotor) + "/" + str(totalDownMotor)),
    #     ("Truck :", str(totalUpTruck) + "/" + str(totalDownTruck)),
    # ]

    # loop over the info tuples and draw them on our frame
    #for (i, (k, v)) in enumerate(info):
        #    text = "{}: {}".format(k, v)
        #    cv2.putText(frame, text, (10, frameHeight - ((i * 60) + 60)),
    #             cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2)

def countingObject2(objects, trackableObjectList, type, frame):
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]

    global totalDown
    global totalUp
    global totalUpCar
    global totalUpBus
    global totalUpMotor
    global totalUpTruck
    global totalDownCar
    global totalDownBus
    global totalDownTruck
    global totalDownMotor
    # loop over the tracked objects
    for (objectID, centroid) in objects.items():
        # check to see if a trackable object exists for the current
        # object ID
        to = trackableObjectList.get(objectID, None)

        # if there is no existing trackable object, create one
        if to is None:
            to = TrackableObject(objectID, centroid)

        # otherwise, there is a trackable object so we can utilize it
        # to determine direction
        else:
            # the difference between the y-coordinate of the *current*
            # centroid and the mean of *previous* centroids will tell
            # us in which direction the object is moving (negative for
            # 'up' and positive for 'down')
            y = [c[1] for c in to.centroids]
            direction = centroid[1] - np.mean(y)
            to.centroids.append(centroid)

            # check to see if the object has been counted or not
            if not to.counted:
                # if the direction is negative (indicating the object
                # is moving up) AND the centroid is above the center
                # line, count the object
                point = Point(centroid[0],centroid[1])

               # if direction < 0 and centroid[1] in range(frameHeight // 2 - 30, frameHeight // 2 + 30):
                if  getOnlyOnSegmentPoints(pointEntree1,pointEntree2,point,1.5):
                    print("Voici le type "+type+" Entrée")
                    print("x" + str(centroid[0]) + " y" + str(centroid[1]))
                    print("Entree1 x" + str(pointEntree1.x) + " y" + str(pointEntree1.y)+"|| x"+str(pointEntree2.x)+" y"+str(pointEntree2.y))

                    #if type is "person":
                     #   totalUp += 1
                    if type is "car":
                        totalUpCar += 1
                    if type is "bus":
                        totalUpBus += 1
                    if type is "motorbike":
                        totalDownMotor += 1
                    if type is "truck":
                        totalUpTruck += 1
                    to.counted = True

                # if the direction is positive (indicating the object
                # is moving down) AND the centroid is below the
                # center line, count the object
                #elif direction > 0 and centroid[1] in range(frameHeight // 2 - 30, frameHeight // 2 + 30)
                elif  getOnlyOnSegmentPoints(pointSortie1,pointSortie2,point,1.5):
                    print("Voici le type " + type+" Sortie")
                    print("x" + str(centroid[0]) + " y" + str(centroid[1]))
                    print("Entree1 x" + str(pointSortie1.x) + " y" + str(pointSortie1.y)+"|| x"+str(pointSortie2.x)+" y"+str(pointSortie2.y))
                   # if type is "person":
                    #    totalDown += 1
                     #   print("OjbectId"+str(objectID))
                    if type is "car":
                        totalDownCar += 1
                        print("OjbectId"+str(objectID))
                    if type is "bus":
                        totalDownBus += 1
                    if type is "motorbike":
                        totalDownMotor += 1
                    if type is "truck":
                        totalDownTruck += 1
                    to.counted = True
            #else:
              #  print("OjbectId" + str(objectID) + " Centroid xy " + str(centroid[0]) + ":" + str(
               #     centroid[1]) + " IsCounted " + str(to.counted))
        # store the trackable object in our dictionary
        trackableObjectList[objectID] = to
        # draw both the ID of the object and the centroid of the
        # object on the output frame
        # text = "ID {}".format(objectID)
        # cv.putText(frame, text, (centroid[0] - 10, centroid[1] - 10),
        # cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
        cv2.circle(frame, (centroid[0], centroid[1]), 4, (0, 255, 0), -1)

def search_video():
    global video
    global b1
    global b2
    global cap
    global tkImage
    global tree
    global duration

    video = filedialog.askopenfilename()
    if(video):
        b1.state(["active"])
        b2["state"] = tk.NORMAL
        b1.state(["active"])

        cap = cv2.VideoCapture(video)

        hasFrame, frame = cap.read()
        frame = cv2.resize(frame, (800, 600))
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        tkImage = ImageTk.PhotoImage(img)
        canvas.create_image(0, 0, anchor=NW, image=tkImage)
        fps = cap.get(cv2.CAP_PROP_FPS)  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = frame_count / fps
        tree.insert("", 'end', text="L1", values=(video, str(datetime.timedelta(seconds=duration)).split(".")[0]))
        #cap = cv2.VideoCapture(video)
        #show_frame()


def search_video2():
    global videos
    global b1
    global b2
    global cap
    global tkImage
    global tree
    global duration
    global videoList
    global video
    global indexVideo
    global videoAttenteList
    indexVideo=0

    videos = filedialog.askopenfilenames()

    if (videos):
        b1.state(["active"])
        b2["state"] = tk.NORMAL
        b1.state(["active"])
        for videoT in videos:


            capT = cv2.VideoCapture(videoT)

            #hasFrame, frame = capT.read()
            #frame = cv2.resize(frame, (800, 600))
            #cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
           # img = Image.fromarray(cv2image)
           # tkImage = ImageTk.PhotoImage(img)
           # canvas.create_image(0, 0, anchor=NW, image=tkImage)
            fps = capT.get(cv2.CAP_PROP_FPS)  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
            frame_count = int(capT.get(cv2.CAP_PROP_FRAME_COUNT))
            durationT = frame_count / fps
            videTmp = Video(videoT, "Non spécifié",str(datetime.timedelta(seconds=durationT)).split(".")[0],"En attente")
            videoList.append(videTmp)
            tree.insert("", 'end', text="L1", values=(videoT,"Non spécifié" ,str(datetime.timedelta(seconds=durationT)).split(".")[0],"En attente"))
        # cap = cv2.VideoCapture(video)
        video=videoList[0].chemin
        capT = cv2.VideoCapture(videoList[0].chemin)

        hasFrame, frame = capT.read()
        frame = cv2.resize(frame, (800, 600))
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        tkImage = ImageTk.PhotoImage(img)
        canvas.create_image(0, 0, anchor=NW, image=tkImage)
        fps = capT.get(cv2.CAP_PROP_FPS)  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
        frame_count = int(capT.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = frame_count / fps
def nextVideo():
    global indexVideo
    global cap
    global tkImage
    global tree
    global duration
    global videoList
    global video
    global indexVideo
    global fps
    global frame_count
    global duration
    if indexVideo + 1 < (len(videoList) ):
        indexVideo+=1
        video = videoList[indexVideo].chemin
        capT = cv2.VideoCapture(videoList[indexVideo].chemin)

        hasFrame, frame = capT.read()
        frame = cv2.resize(frame, (800, 600))
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        tkImage = ImageTk.PhotoImage(img)
        canvas.create_image(0, 0, anchor=NW, image=tkImage)

        fps = capT.get(cv2.CAP_PROP_FPS)  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
        frame_count = int(capT.get(cv2.CAP_PROP_FRAME_COUNT))
        durationT = frame_count / fps

       # canvas.pack_forget()

def prevVideo():
    global indexVideo

    global tkImage
    global tree
    global duration
    global videoList
    global video
    global indexVideo

    if indexVideo - 1 >= 0:
        indexVideo-=1
        video = videoList[indexVideo].chemin
        capT = cv2.VideoCapture(videoList[indexVideo].chemin)

        hasFrame, frame = capT.read()
        frame = cv2.resize(frame, (800, 600))
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        tkImage = ImageTk.PhotoImage(img)
        canvas.create_image(0, 0, anchor=NW, image=tkImage)
        fps = capT.get(cv2.CAP_PROP_FPS)  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
        frame_count = int(capT.get(cv2.CAP_PROP_FRAME_COUNT))
        durationT = frame_count / fps


def insertData():
    global historiqueList
    histoService=HistoriqueComtpageService()
    histo=Historique_comptage(0,"D:/Stage/67 HA/D2/DJI_0007.MP4","2000-12-02","0:07:00","67HA SUD",12,22,6,7,30,17)
    histoService.insertIntoTable(histo)
    #tree.insert("", 'end', text="L1",values=(, "Non spécifié", str(datetime.timedelta(seconds=duration)).split(".")[0], "En attente"))
    refreshHisto()
def insertData2():
    global historiqueList
    histoService=HistoriqueComtpageService()
    histo=Historique_comptage(0,"D:/Stage/67 HA/D2/DJI_0008.MP4","2000-12-02","0:09:00","HAVANA",9,15,5,0,10,19)
    histoService.insertIntoTable(histo)
    #tree.insert("", 'end', text="L1",values=(, "Non spécifié", str(datetime.timedelta(seconds=duration)).split(".")[0], "En attente"))
    refreshHisto()
def insertHisto(comptage):
    global historiqueList
    histoService=HistoriqueComtpageService()
    #histo=Historique_comptage(0,"Chemin","2000-12-02","0:07:00","67HA",7,8,5,9,6)
    histoService.insertIntoTable(comptage)
    #tree.insert("", 'end', text="L1",values=(, "Non spécifié", str(datetime.timedelta(seconds=duration)).split(".")[0], "En attente"))
    refreshHisto()

def test_program_thread():
    thread = threading.Thread(None, edit_item, None, (), {})
    thread.start()



def edit_item():
    global tree
    focused = tree.focus()
    x = input('Enter a Value you want to change')
    tree.insert("", str(focused)[1:], values=("", str(x)))
    tree.delete(focused)
def popup_bonus():
    win = tk.Toplevel()
    win.wm_title("Window")
    win.geometry("400x50")
    windowWidth = win.winfo_reqwidth()
    windowHeight = win.winfo_reqheight()

    # Gets both half the screen width/height and window width/height
    positionRight = int(win.winfo_screenwidth() / 2 - windowWidth / 2)
    positionDown = int(win.winfo_screenheight() / 2 - windowHeight / 2)
    checkImg = PhotoImage(file="checked.png")
    # Positions the window in the center of the page.
    win.geometry("+{}+{}".format(positionRight, positionDown))
    l2 = tk.Label(win, image=checkImg)
    l2.place(x=10, y=10)
    l = tk.Label(win, text="Comptage terminé, Veuillez retracer les lignes de comptage")
    l.place(x=50,y=10)


def treeview_sort_column(tv, col, reverse):
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    l.sort(reverse=reverse)

    # rearrange items in sorted positions
    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)

    # reverse sort next time
    tv.heading(col, command=lambda:treeview_sort_column(tv, col, not reverse))


def select_files():
    filetypes = (
        ('mp4', '*.mp4'),
        ('All files', '*.*')
    )

    filenames = filedialog.askopenfilenames(
        title='Open files',
        initialdir='/',
        filetypes=filetypes)


    showinfo(
        title='Selected Files',
        message=filenames[0]
    )
def detect_objectsYolo():
    global videoList
    global indexVideo
    global cap
    global duration
    global display1
    while cap.isOpened():
        global pointEntree1
        global pointEntree2
        global pointSortie1
        global pointSortie2
        global fps

        hasFrame, frame = cap.read()
        if hasFrame is False:
            videoTmp=videoList[indexVideo]
            #sheet["C2"] = totalUpCar + totalDownCar
            #sheet["D2"] = totalUpBus + totalDownBus
            #sheet["E2"] = totalUpTruck + totalDownTruck

            comptageHisto=Historique_comptage(0,videoList[indexVideo].chemin,"2000-20-12",videoList[indexVideo].duree,videoList[indexVideo].lieu,totalUpCar + totalDownCar,totalUpBus + totalDownBus,totalUpTruck + totalDownTruck,totalDownMotor+totalUpMotor,totalDownCar + totalDownTruck + totalDownBus,totalUpCar + totalUpTruck + totalUpBus)
            insertHisto(comptageHisto)

            break
        frame = cv2.resize(frame, (800, 600), interpolation=cv2.INTER_AREA)



        cv2.line(frame, (pointEntree1.x, pointEntree1.y), (pointEntree2.x, pointEntree2.y), (52, 235, 58), 2)
        cv2.line(frame, (pointSortie1.x, pointSortie1.y), (pointSortie2.x, pointSortie2.y), (0, 47, 255), 2)
        # Create a 4D blob from a frame.
        blob = cv2.dnn.blobFromImage(frame, 1 / 255, (inpWidth, inpHeight), [0, 0, 0], 1, crop=False)

        # Sets the input to the network
        net.setInput(blob)

        # Runs the forward pass to get output of the output layers
        outs = net.forward(getOutputsNames(net))

        # Remove the bounding boxes with low confidence
        postprocess2(frame, outs)

        t, _ = net.getPerfProfile()

        # show the output image
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        tkimage = ImageTk.PhotoImage(img)
        display1.imgtk = tkimage  # Shows frame for display 1
        display1.configure(image=tkimage)
        # boucle()
        # threadDetection()
        display1.update()
        #display1.after(1, detect_objectsYolo)
    canvas.pack()
    threadNext()
    threadP()


        #detect_objectsYolo()
def initUser():
    userService = UserService()
    user = User(0,"root","root","thongsangchad@gmail.com")
    userService.insertIntoTable(user)
    # tree.insert("", 'end', text="L1",values=(, "Non spécifié", str(datetime.timedelta(seconds=duration)).split(".")[0], "En attente"))
    refreshProfile()
def threadNext():
    # Call work function
    t1 = Thread(target=nextVideo())
    t1.start()
def threadP():
    # Call work function
    t1 = Thread(target=popup_bonus())
    t1.start()

def threadN():
    # Call work function
    t1 = Thread(target=threadExcel())
    t1.start()
def threadExcel():
    # Call work function
    t1 = Thread(target=ExportExcel())
    t1.start()

def threadSearch():
    t2 = Thread(target=search_video())
    t2.start()

def threadPlay():
    t2 = Thread(target=Play())
    t2.start()
def threadPause():
    t2 = Thread(target=Pause())
    t2.start()
def boucle():
    te = Thread(target=detect_objectsYolo())
    te.start()

def Pause():
    display1.after_cancel(detect_objectsYolo)
def ExportExcel():
    files = [('Excel Files', '*.xlsx')]
    file = filedialog.asksaveasfile(filetypes=files, defaultextension=files)

    workbook = Workbook()
    sheet = workbook.active
    totalEntree = 0
    totalSortie = 0
    sheet["A1"] = "LIEU"
    sheet["B1"] = "VIDEO"
    sheet["C1"] = "CAR"
    sheet["D1"] = "BUS"
    sheet["E1"] = "TRUCK"
    sheet["F1"] = "DUREE CAPTURE"
    sheet["G1"] = "ENTREE"
    sheet["H1"] = "1H"
    sheet["I1"] = "14H"
    sheet["J1"] = "SORTIE"
    sheet["K1"] = "1H"
    sheet["L1"] = "14H"
    sheet["M1"] = "TOTAL TRAFFIC"

    sheet["A2"] = "LIEU"
    sheet["B2"] = video
    sheet["C2"] = totalUpCar + totalDownCar
    sheet["D2"] = totalUpBus + totalDownBus
    sheet["E2"] = totalUpTruck + totalDownTruck
    sheet["F2"] = int(duration / 60)
    sheet["G2"] = totalDownCar + totalDownTruck + totalDownBus
    if totalDownCar != 0 or totalDownTruck != 0 or totalDownBus != 0:
        totalEntree = int(((totalDownCar + totalDownTruck + totalDownBus) * 840) / int(duration / 60))
        sheet["H2"] = int(((totalDownCar + totalDownTruck + totalDownBus) * 60) / int(duration / 60))
        sheet["I2"] = totalEntree

    else:
        sheet["H2"] = 0
        sheet["I2"] = 0
    sheet["J2"] = totalUpCar + totalUpTruck + totalUpBus
    if totalUpCar != 0 or totalUpTruck != 0 or totalUpBus != 0:
        totalSortie = int(((totalUpCar + totalUpTruck + totalUpBus) * 840) / int(duration / 60))
        sheet["K2"] = int(((totalUpCar + totalUpTruck + totalUpBus) * 60) / int(duration / 60))
        sheet["L2"] = totalSortie

    else:
        sheet["K2"] = 0
        sheet["L2"] = 0
    ##FFFF00 jaune
    redFill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00',
                          fill_type='solid')
    sheet["M2"].fill = redFill
    sheet["M2"] = totalSortie + totalEntree

    workbook.save(filename=file.name)

def ExcelExport():
    global historiqueList
    global histoDataList
    files = [('Excel Files', '*.xls')]
    file = filedialog.asksaveasfile(filetypes=files, defaultextension=files)
    workbook = xlwt.Workbook()
    excelSheet = workbook.add_sheet("Resultat")
    #print("taille",str(len(historiqueList.get_children())))
    var=vars(histoDataList[0])
    print("taille var",len(var))
    indice=0
    tmp=0
    for index in range(len(var)) :
        print(index)
        excelSheet.write(0, index, historiqueList.heading(index+1)["text"])
    for i, j in enumerate(historiqueList.get_children()):
        values = historiqueList.item(j)["values"]
        indice+=1
        for k, m in enumerate(values):
                excelSheet.write(indice, k,m)
    workbook.save(file.name)
def CSVExport():
    global historiqueList
    global histoDataList
    files = [('CSV Files', '*.csv')]
    file = filedialog.asksaveasfile(filetypes=files, defaultextension=files)

    with open(file.name, mode='w') as csv_file:
        #writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        csvwriter = csv.writer(csv_file, delimiter=',')

        for row_id in historiqueList.get_children():
            row = historiqueList.item(row_id)['values']
            print('save row:', row)
            csvwriter.writerow(row)

def createNewWindow():
    newWindow = tk.Toplevel(window)

    labelExample = tk.Label(newWindow, text="New Window")
    buttonExample = Button(newWindow, text="Exporter Excel", command=threadExcel)

    labelExample.pack()
    buttonExample.pack()


# Set up GUI
window = ThemedTk(theme="adapta")  # Makes main window
window.wm_title("Comptage de vehicule")
window.config(background="#FFFFFF")
window.resizable(True, True)
window.geometry("1366x800")

#imageFrame.pack(side=tk.LEFT,pady=5)

#window.config(menu=menubar)

notebook = ttk.Notebook(window)
notebook.pack(expand=True)

# create frames
menuFrame = ttk.Frame(notebook, width=1360, height=800)
historiqueFrame = ttk.Frame(notebook, width=1360, height=800)
profilFrame = ttk.Frame(notebook, width=1360, height=800)

menuFrame.pack(fill='both', expand=True)
historiqueFrame.pack(fill='both', expand=True)
#profilFrame.pack(fill='both', expand=True)

# add frames to notebook

notebook.add(menuFrame, text='Comptage')
notebook.add(historiqueFrame, text='Historique')
#notebook.add(profilFrame, text='Profil')

# Capture video frames

# cap = cv2.VideoCapture("video.mp4")
def print_selection():
    global showCar
    global showPerson
    global showTruck
    global showMotorBike
    global showBus
    global isEntree
    global isSortie
    global  canDrawOut
    global canDrawEnter
    global c6
    global c7
    if (var1.get() == 1):
        showPerson=True
    elif (var1.get() == 0):
        showPerson=False
    if (var2.get() == 1):
        showCar=True
    elif (var2.get() == 0):
        showCar=False
    if (var3.get() == 1):
        showTruck=True
    elif (var3.get() == 0):
        showTruck=False
    if (var4.get() == 1):
        showBus=True
    elif (var4.get() == 0):
        showBus=False
    if (var5.get() == 1):
        showMotorBike=True
    elif (var5.get() == 0):
        showMotorBike=False
    if (var6.get() == 0):
        isEntree=False

    elif (var6.get() == 1):
        isEntree=True
        canDrawEnter = 1
        isSortie=False
        canDrawOut=0


    if (var7.get() == 0):
        isSortie = False
    elif (var7.get() == 1):
        isSortie = True
        canDrawOut = 1
        isEntree=False
        canDrawEnter=0
       # c6.deselect()





def RefreshLabel():
    global  label4
    global label5
    global label6
    global label7
    global label8
    global label9
    global label10
    global label11
    global label12
    global label13
    global label15
    global label16
    global fps
    global duration
    # valeurs
    label4 = ttk.Label(canvas1, text=totalUpCar)
    label5 = ttk.Label(canvas1, text=totalDownCar)
    label6 = ttk.Label(canvas1, text=totalUpTruck)
    label7 = ttk.Label(canvas1, text=totalDownTruck)
    label8 = ttk.Label(canvas1, text=totalUpBus)
    label9 = ttk.Label(canvas1, text=totalDownBus)
    label10 = ttk.Label(canvas1, text=totalUpMotor)
    label11 = ttk.Label(canvas1, text=totalDownMotor)
    #label15 = Label(window, text="Durée : "+str(datetime.timedelta(seconds=duration)).split(".")[0], bg="white")


    label1.place(x=10, y=10)
    label2.place(x=100, y=10)
    label3.place(x=150, y=10)

    label4.place(x=100, y=40)
    label5.place(x=150, y=40)
    label6.place(x=100, y=70)
    label7.place(x=150, y=70)
    label8.place(x=100, y=100)
    label9.place(x=150, y=100)
    label10.place(x=100, y=130)
    label11.place(x=150, y=130)
    #label15.place(x=860, y=15)


    canvas1.after(1000, RefreshLabel)
def refreshProfile():
    global labelLogin
    global labelEmail
    global labelPass
    global userProfile
    userService=UserService()
    userProfile=userService.readAll()[0]
    labelLogin = ttk.Label(profilFrame, text=userProfile.login)
    labelLogin.place(x=100, y=40)

    labelPass = ttk.Label(profilFrame, text=userProfile.mdp)
    labelPass.place(x=130, y=60)

    labelEmail = ttk.Label(profilFrame, text=userProfile.email)
    labelEmail.place(x=100, y=80)


def refreshAttente():
    global videoList

def refreshHisto():
    global historiqueList
    global histoDataList
    histoDataList=[]
    for item in historiqueList.get_children():
        historiqueList.delete(item)
    histoService=HistoriqueComtpageService()
    results=histoService.readAll()
    for res in results:
        #(self, id, nom, date, duree, lieu, nbCar, nbBus, nbTruck, entree, sortie)
        histoDataList.append(res)
        historiqueList.insert("", 'end', text="L1",values=(res.id,res.nom, res.date,res.duree,res.lieu,res.nbCar,res.nbBus,res.nbTruck,res.nbMotorBike,res.entree,res.sortie))
def findHisto():

    global historiqueList
    global histoDataList
    global entryLieu
    global dateEntry1
    global dateEntry2
    #compt=Historique_comptage(0,"",str(dateEntry1.get_date()),"7",entryLieu.get(),21,2,3,4,4,7)
    critere=Critere(entryLieu.get(),str(dateEntry1.get_date()),str(dateEntry2.get_date()))
    histoDataList=[]
    for item in historiqueList.get_children():
        historiqueList.delete(item)
    histoService=HistoriqueComtpageService()
    results=histoService.findByCriteria(critere)
    for res in results:
        #(self, id, nom, date, duree, lieu, nbCar, nbBus, nbTruck, entree, sortie)
        histoDataList.append(res)
        historiqueList.insert("", 'end', text="L1",values=(res.id,res.nom, res.date,res.duree,res.lieu,res.nbCar,res.nbBus,res.nbTruck,res.nbMotorBike,res.entree,res.sortie))

def draw(event):
    x, y = event.x, event.y
    if canvas.old_coords:
        x1, y1 = canvas.old_coords
        canvas.create_line(x, y, x1, y1)
    canvas.old_coords = x, y

def draw_line(event):
    global canDrawEnter
    global canDrawOut
    global  isEntree
    global isSortie
    global canvas
    global pointEntree1
    global pointEntree2
    global pointSortie1
    global pointSortie2
    global c6
    global c7

    if str(event.type) == 'ButtonPress':
        if(isEntree or isSortie):
            canvas.old_coords = event.x, event.y




    elif str(event.type) == 'ButtonRelease':
        if(canvas.old_coords):
            x, y = event.x, event.y
            x1, y1 = canvas.old_coords
            print("CanDrawEnter ",canDrawEnter)
            print("isEnter ", canDrawOut)
            print("isSortie ", isSortie)
            if(isEntree and canDrawEnter==1):
                canvas.create_line(x, y, x1, y1,fill='green')
                pointEntree1 = Point(x1, y1)
                canDrawEnter=0
                print_selection()
                pointEntree2=Point(x,y)
            if (isSortie and canDrawOut==1):
                canvas.create_line(x, y, x1, y1,fill='red')
                pointSortie1 = Point(x1, y1)

                canDrawOut = 0
                print_selection()
                pointSortie2 = Point(x, y)
            reset_coords(event)

def reset_coords(event):
    canvas.old_coords = None
    global pointSortie2

def ClearCanvas():
    global canDrawEnter
    global canDrawOut
    global canvas
    global videoList
    global indexVideo
    canDrawEnter = 1
    canDrawOut = 1
    canvas.delete('all')
    capT = cv2.VideoCapture(videoList[indexVideo].chemin)

    hasFrame, frame = capT.read()
    frame = cv2.resize(frame, (800, 600))
    cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
    img = Image.fromarray(cv2image)
    tkImage = ImageTk.PhotoImage(img)
    canvas.create_image(0, 0, anchor=NW, image=tkImage)
    canvas.update()
    print("Is Enter ato am clear", isEntree)
    print("Can Draw Out ato am clear", canDrawEnter)


def Play():
    global  video
    global videoList
    global indexVideo
    if videoList[indexVideo].chemin:
        global cap
        global fps
        global frame_count
        global duration
        canvas.pack_forget()
        cap = cv2.VideoCapture(videoList[indexVideo].chemin)
        fps = cap.get(cv2.CAP_PROP_FPS)  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = frame_count / fps
        detect_objectsYolo()
def show_frame():
    if cap.isOpened():
        _, frame = cap.read()
        frame = cv2.flip(frame, 1)
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        imgtk = ImageTk.PhotoImage(image=img)
        display1.imgtk = imgtk  # Shows frame for display 1
        display1.configure(image=imgtk)
        display1.update()
        window.update()
   # window.after(10, show_frame)
def showChart(event):
    global historiqueList
    global videoList
    global stockListExp
    global fig
    global ax
    global chart1
    global historiqueFrame
    global histoDataList
    print("Selected", historiqueList.index(historiqueList.selection()))

    indice=int(historiqueList.index(historiqueList.selection()))
    print("Indice", indice)
    total=int(histoDataList[indice].entree)+int(histoDataList[indice].sortie)
    print("Indice", indice)
    stockListExp = ['CAR' , 'BUS', 'TRUCK', 'MOTORBIKE']
    if total!=0:
        stockSplitExp = [int(histoDataList[indice].nbCar*100)/total,int(histoDataList[indice].nbBus*100)/total,int(histoDataList[indice].nbTruck*100)/total,int(histoDataList[indice].nbMotorBike*100)/total]
    else:
        stockSplitExp=[0,0,0,0]
    print(stockSplitExp)
    fig = Figure(figsize=(3,3)) # create a figure object
    ax = fig.add_subplot(111) # add an Axes to the figure
    fig.patch.set_facecolor('white')
    ax.pie(stockSplitExp, radius=0.5, labels=stockListExp,autopct='%0.0f%%', shadow=False)

    chart1 = FigureCanvasTkAgg(fig,historiqueFrame)
    chart1.draw()
    chart1.get_tk_widget().place(x=1040,y=10)
def display_msg():
    global dateEntry

    date = dateEntry.get_date()

    print(date)
    t = f"Your appointment is booked for {date} "
def showDate():
    global labelLieu
    datePicker = MyDatePicker(labelLieu,format_str='%02d-%s-%s')
    date=datePicker.dateFull
    print("ITO FULL DATE",date)
def set_cell_value(event):
    global tree
    global  frame1
    global videoList
    global item_text
    for index,item in enumerate(tree.selection()):
        item_text = tree.item(item, "values")
        column = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
    cn = int(str(column).replace('#', ''))
    rn = int(str(row).replace('I', ''))
    print(rn)
    entryedit = Text(frame1, width=5 + (cn - 1) * 16, height=1)
    entryedit.insert(INSERT,item_text[cn-1])
    entryedit.place(x= (cn - 1) * 130, y=6 + rn * 60)
    val=rn-1
    def saveedit():
        global item_text
        tree.set(item, column=column, value=entryedit.get(0.0, "end"))
        print("Cn"+str(cn))
        print("Rn"+str(rn))
        print("Val"+str(val))

        if cn == 1:
            videoList[val].chemin=entryedit.get(0.0, "end").strip()
        if cn == 2:
            videoList[val].lieu=entryedit.get(0.0, "end").strip()
        if cn == 3:
            videoList[val].duree=entryedit.get(0.0, "end").strip()
        if cn == 4:
            videoList[val].etat=entryedit.get(0.0, "end").strip()
        print(videoList[val].lieu)
        entryedit.destroy()
        okb.destroy()

    okb = ttk.Button(frame1, text='OK', width=4, command=saveedit)
    okb.place(x=90 + (cn - 1) * 242, y=2 + rn * 60)


canvas1 = ttk.Frame(menuFrame, width=250, height=500)
canvas1.place(x=0,y=10)


label1 = ttk.Label(canvas1, text="Type ")
label2 = ttk.Label(canvas1, text="Entrée ")
label3 = ttk.Label(canvas1, text="Sortie")

#valeurs
#label4 = ttk.Label(canvas1, text=totalUp)
#label5 = ttk.Label(canvas1, text=totalDown)
label4 = ttk.Label(canvas1, text=totalUpCar)
label5 = ttk.Label(canvas1, text=totalDownCar)
label6 = ttk.Label(canvas1, text=totalUpTruck)
label7 = ttk.Label(canvas1, text=totalDownTruck)
label8 = ttk.Label(canvas1, text=totalUpBus)
label9 = ttk.Label(canvas1, text=totalDownBus)
label10 = ttk.Label(canvas1, text=totalUpMotor)
label11 = ttk.Label(canvas1, text=totalDownMotor)
label12 = ttk.Label(canvas1, text=totalUpMotor)
label13 = ttk.Label(canvas1, text=totalDownMotor)
label14 = ttk.Label(canvas1, text="Tracage des lignes")
label15 = ttk.Label(window, text="Durée 00:00:00")


label1.place(x=10,y=10)
label2.place(x=100,y=10)
label3.place(x=150,y=10)

label4.place(x=100,y=40)
label5.place(x=150,y=40)
label6.place(x=100,y=70)
label7.place(x=150,y=70)
label8.place(x=100,y=100)
label9.place(x=150,y=100)
label10.place(x=100,y=130)
label11.place(x=150,y=130)
#label12.place(x=100,y=160)
#label13.place(x=150,y=160)
label14.place(x=10,y=210)
#label15.place(x=860,y=15)


# Graphics window

imageFrame = tk.Frame(menuFrame, width=600, height=500)
imageFrame.place(x=250,y=10)

display1 = tk.Label(imageFrame, width=600, height=500)
display1.place(x=0,y=0)




canvas = tk.Canvas(imageFrame, width=600, height=500)

canvas.place(x=0,y=0)


canvas.pack()

canvas.old_coords = None


var1 = tk.IntVar()
var2 = tk.IntVar()
var3 = tk.IntVar()
var4 = tk.IntVar()
var5 = tk.IntVar()
var6 = tk.IntVar()
var7 = tk.IntVar()
c1 = tk.Checkbutton(canvas1, text='Person', variable=var1, onvalue=1, offvalue=0, command=print_selection,fg="red",bg="white")

c2 = tk.Checkbutton(canvas1, text='Voiture', variable=var2, onvalue=1, offvalue=0, command=print_selection,fg="blue",bg="white")
c3 = tk.Checkbutton(canvas1, text='Camion', variable=var3, onvalue=1, offvalue=0, command=print_selection,fg="green",bg="white")
c4 = tk.Checkbutton(canvas1, text='Bus', variable=var4, onvalue=1, offvalue=0, command=print_selection,fg="yellow",bg="white")
c5 = tk.Checkbutton(canvas1, text='Moto', variable=var5, onvalue=1, offvalue=0, command=print_selection,fg="orange",bg="white")
s = ttk.Style()
s.configure('B1.TCheckbutton')
s.map('B1.TCheckbutton',
        foreground=[('disabled', 'white'),
                    ('pressed', 'green'),('selected', 'green'),
                    ('active', 'green')],

      )

c6 = ttk.Checkbutton(canvas1,style='B1.TCheckbutton', text='Ligne Entrée', variable=var6, onvalue=1, offvalue=0, command=print_selection)
s2 = ttk.Style()
s2.configure('B2.TCheckbutton')
s.map('B2.TCheckbutton',
        foreground=[('disabled', 'grey'),
                    ('pressed', 'red'),('selected', 'red'),
                    ('active', 'red')])
c7 = ttk.Checkbutton(canvas1,style='B2.TCheckbutton', text='Ligne Sortie', variable=var7, onvalue=1, offvalue=0, command=print_selection)
#c7 = tk.Checkbutton(canvas1, text='Ligne Sortie', variable=var7, onvalue=1, offvalue=0, command=print_selection,bg="white",fg="red")

#c1.place(x=10,y=40)
#c2.place(x=10,y=70)
c2.place(x=10,y=40)
c3.place(x=10,y=70)
c4.place(x=10,y=100)
c5.place(x=10,y=130)
c6.place(x=10,y=240)
c7.place(x=10,y=270)
###LISTE DE LA FILE D ATTENTE########
frame1 = ttk.Frame(menuFrame, width=600, height=500)
frame1.place(x=900,y=10)
tree = ttk.Treeview(frame1, selectmode='browse')
tree.place(x=10, y=40)

vsb = ttk.Scrollbar(frame1, orient="vertical", command=tree.yview)
vsb.place(x=10+400+2, y=40, height=200+20)

tree.configure(yscrollcommand=vsb.set)

tree["columns"] = ("1", "2", "3","4")
tree['show'] = 'headings'
tree.column("1", width=180, anchor='c')
tree.column("2", width=60, anchor='c')
tree.column("3", width=60, anchor='c')
tree.column("4", width=60, anchor='c')
tree.heading("1", text="Nom")
tree.heading("2", text="Lieu")
tree.heading("3", text="Durée")
tree.heading("4", text="Etat")
tree.bind('<Double-1>', set_cell_value)

#######LISTE DES HISTORIQUES DE COMPTAGE###########
frameHist = ttk.Frame(historiqueFrame, width=1300, height=500)
frameHist.place(x=0,y=10)
historiqueList = ttk.Treeview(frameHist, selectmode='browse')
historiqueList.place(x=10, y=40)

vsb = ttk.Scrollbar(frameHist, orient="vertical", command=historiqueList.yview)
vsb.place(x=10+900+2, y=40, height=200+20)

historiqueList.configure(yscrollcommand=vsb.set)

historiqueList["columns"] = ("1", "2", "3","4","5","6","7","8","9","10","11")
historiqueList['show'] = 'headings'
historiqueList.column("1", width=50, anchor='c')
historiqueList.column("2", width=100, anchor='c')
historiqueList.column("3", width=100, anchor='c')
historiqueList.column("4", width=100, anchor='c')
historiqueList.column("5", width=70, anchor='c')
historiqueList.column("6", width=70, anchor='c')
historiqueList.column("7", width=70, anchor='c')
historiqueList.column("8", width=70, anchor='c')
historiqueList.column("9", width=70, anchor='c')
historiqueList.column("10", width=70, anchor='c')
historiqueList.column("11", width=70, anchor='c')
historiqueList.heading("1", text="N°")
historiqueList.heading("2", text="Nom")
historiqueList.heading("3", text="Date", command=lambda:treeview_sort_column(historiqueList, "3", False))
historiqueList.heading("4", text="Durée", command=lambda:treeview_sort_column(historiqueList, "4", False))
historiqueList.heading("5", text="Lieu", command=lambda:treeview_sort_column(historiqueList, "5", False))
historiqueList.heading("6", text="Voiture", command=lambda:treeview_sort_column(historiqueList, "6", False))
historiqueList.heading("7", text="Bus", command=lambda:treeview_sort_column(historiqueList, "7", False))
historiqueList.heading("8", text="Camion", command=lambda:treeview_sort_column(historiqueList, "8", False))
historiqueList.heading("9", text="Moto", command=lambda:treeview_sort_column(historiqueList, "9", False))
historiqueList.heading("10", text="TotalEntree", command=lambda:treeview_sort_column(historiqueList, "10", False))
historiqueList.heading("11", text="TotalSortie", command=lambda:treeview_sort_column(historiqueList, "11", False))
historiqueList.bind('<Button-1>', showChart)
refreshHisto()
################################Liste comptage en cours#########################


################################################


entry6_img = PhotoImage(file = "play-button.png")
myFont = font.Font(family='Helvetica', size=10, weight='bold')
b1 = ttk.Button(frame1,image=entry6_img,text = "Lancer",command = threadPlay)

#b1.place(x=190,y=300)
b1.place(x=200,y=0)
#b1['font']=myFont




b2 = ttk.Button(canvas1,text = "Exporter les résultats ",command = threadExcel)
#b2.place(x=10,y=400)


b3 = ttk.Button(canvas1,text = "Effacer les lignes",command = ClearCanvas)
b3.place(x=10,y=350)

entry1_img = PhotoImage(file = "plus.png")
entry2_img = PhotoImage(file = "right-arrow.png")
entry3_img = PhotoImage(file = "left-arrow.png")
entry4_img = PhotoImage(file = "excel.png")
entry5_img = PhotoImage(file = "csv.png")
ajoutBtn = ttk.Button(frame1,image=entry1_img, compound="left",text = "Ajouter à la file d'attente",command = search_video2)
ajoutBtn.place(x=10,y=0)

nextBtn = ttk.Button(menuFrame,image=entry2_img, compound="left",command = nextVideo)
nextBtn.place(x=852,y=220,width=50)
prevBtn = ttk.Button(menuFrame,image=entry3_img, compound="left",command = prevVideo)
prevBtn.place(x=200,y=220,width=50)

exportExcelBtn = ttk.Button(historiqueFrame,image=entry4_img, compound="left",text = "Exporter en Excel",command = ExcelExport)
exportExcelBtn.place(x=10,y=0)
exportCSVBtn = ttk.Button(historiqueFrame,image=entry5_img, compound="left",text = "Exporter en CSV",command = CSVExport)
exportCSVBtn.place(x=200,y=0)

labelLieu = ttk.Label(historiqueFrame, text="Lieu")
labelLieu.place(x=50,y=360)
labelDate = ttk.Label(historiqueFrame, text="Date Debut")
labelDate.place(x=200,y=360)
labelDate2 = ttk.Label(historiqueFrame, text="Date Fin")
labelDate2.place(x=400,y=360)
entryLieu = Entry(historiqueFrame,
    bd = 0,
    bg = "#ececec",
    highlightthickness = 0)

entryLieu.place(
    x = 50, y = 400,
    width = 100,
    height = 20)

rechercheBtn = ttk.Button(historiqueFrame, compound="left",text = "Rechercher",command = findHisto)
rechercheBtn.place(x=600,y=380)

f = ('Times', 20)




cal = Calendar(
    historiqueFrame,
    selectmode="day",
    year=2021,
    month=9,
    day=30
)


dateEntry1=DateEntry(historiqueFrame, locale='fr_FR', date_pattern='dd-mm-yyyy')  # custom formatting
dateEntry1.place(x=200,y=380)


dateEntry2=DateEntry(historiqueFrame, locale='fr_FR', date_pattern='dd-mm-yyyy')  # custom formatting
dateEntry2.place(x=400,y=380)
######Profil##########
labelUtil = ttk.Label(profilFrame, text="Vos détails")
labelUtil.place(x=10,y=10)
labelL = ttk.Label(profilFrame, text="Votre Login:")
labelL.place(x=10,y=40)

labelM = ttk.Label(profilFrame, text="Votre mot de passe:")
labelM.place(x=10,y=60)

labelE = ttk.Label(profilFrame, text="Votre email:")
labelE.place(x=10,y=80)

#initUser()
#refreshProfile()
canvas.bind('<ButtonPress-1>', draw_line)
canvas.bind('<ButtonRelease-1>', draw_line)
c1.select()
c2.select()
c3.select()
c4.select()
c5.select()
print("Compteur")
RefreshLabel()
# show_frame() #Display
window.mainloop()  # Starts GUI


