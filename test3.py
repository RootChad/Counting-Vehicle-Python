import threading
from threading import Thread
from time import sleep

import numpy as np
import numpy as np
import argparse
import cv2
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from PIL import Image, ImageTk
from openpyxl import Workbook
import sys
import os.path
import math

from openpyxl.styles import PatternFill

from counting.centroidtracker import CentroidTracker
from counting.trackableobject import TrackableObject

# Initialize the parameters
confThreshold = 0.6  # Confidence threshold
nmsThreshold = 0.4  # Non-maximum suppression threshold
inpWidth = 416  # Width of network's input image
inpHeight = 416  # Height of network's input image

print("[INFO] loading model...")
# initialize the frame dimensions (we'll set them as soon as we read
# the first frame from the video)
W = None
H = None

# instantiate our centroid tracker, then initialize a list to store
# each of our dlib correlation trackers, followed by a dictionary to
# map each unique object ID to a TrackableObject
ct = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct2 = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct3 = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct4 = CentroidTracker(maxDisappeared=40, maxDistance=50)
ct5 = CentroidTracker(maxDisappeared=40, maxDistance=50)
trackers = []
trackableObjects = {}
trackableObjects2 = {}
trackableObjects3 = {}
trackableObjects4 = {}
trackableObjects5 = {}
# initialize the total number of frames processed thus far, along
# with the total number of objects that have moved either up or down
totalDown = 0
totalUp = 0
totalDownCar = 0
totalUpCar = 0
totalDownBus = 0
totalUpBus = 0
totalDownTruck = 0
totalUpTruck = 0
totalDownMotor = 0
totalUpMotor = 0
# cap = cv2.VideoCapture("video.mp4")
# construct the argument parse and parse the arguments
ap = argparse.ArgumentParser()

ap.add_argument('-l', '--labels', type=str, default='model/coco.names', help='Path to label file')
# ap.add_argument('--video', help='Path to video file.', default='D:/Stage/67 HA/D2/DJI_0004.MP4')

args = ap.parse_args()

# initialize the list of class labels MobileNet SSD was trained to
# detect, then generate a set of bounding box colors for each class
CLASSES = ["background", "aeroplane", "bicycle", "bird", "boat",
           "bottle", "bus", "car", "cat", "chair", "cow", "diningtable",
           "dog", "horse", "motorbike", "person", "pottedplant", "sheep",
           "sofa", "train", "tvmonitor"]
classesFile = open(args.labels).read().strip().split('\n')

# Give the configuration and weight files for the model and load the network using them.
modelConfiguration = "model/yolov3.cfg";
modelWeights = "model/yolov3.weights";

# load our serialized model from disk
print("[INFO] loading model...")
net = cv2.dnn.readNetFromDarknet(modelConfiguration, modelWeights)
net.setPreferableBackend(cv2.dnn.DNN_BACKEND_OPENCV)
net.setPreferableTarget(cv2.dnn.DNN_TARGET_OPENCL)
COLORS = np.random.uniform(0, 255, size=(len(CLASSES), 3))
video = ""
cap = ""
fps = 0
frame_count = 0
duration = 0


def getOutputsNames(net):
    # Get the names of all the layers in the network
    layersNames = net.getLayerNames()
    # Get the names of the output layers, i.e. the layers with unconnected outputs
    return [layersNames[i[0] - 1] for i in net.getUnconnectedOutLayers()]


def draw_bounding_boxes(image, boxes, confidences, classIDs, idxs, colors):
    if len(idxs) > 0:
        for i in idxs.flatten():
            # extract bounding box coordinates
            x, y = boxes[i][0], boxes[i][1]
            w, h = boxes[i][2], boxes[i][3]

            # draw the bounding box and label on the image
            color = [int(c) for c in colors[classIDs[i]]]
            cv2.rectangle(image, (x, y), (x + w, y + h), color, 2)

            text = "{}: {:.4f}".format(classesFile[classIDs[i]], confidences[i])
            cv2.putText(image, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

    return image


# Draw the predicted bounding box
def drawPred(classId, conf, left, top, right, bottom, frame):
    # Draw a bounding box.
    cv2.rectangle(frame, (left, top), (right, bottom), (255, 178, 50), 2)
    # Draw a center of a bounding box
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]
    cv2.line(frame, (0, frameHeight // 2 - 50), (frameWidth, frameHeight // 2 - 50), (0, 255, 255), 2)
    cv2.circle(frame, (left + (right - left) // 2, top + (bottom - top) // 2), 3, (0, 0, 255), -1)

    counter = []
    coun = 0
    if (top + (bottom - top) // 2 in range(frameHeight // 2 - 2, frameHeight // 2 + 2)):
        coun += 1
        # print(coun)

        counter.append(coun)
    # label = 'Pedestrians: '.format(str(counter))
    text = "{}".format(classesFile[classId])
    cv2.putText(frame, text, (left, top - 5), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)


def postprocess2(frame, outs):
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]
    print("ato")
    rects = []

    # Scan through all the bounding boxes output from the network and keep only the
    # ones with high confidence scores. Assign the box's class label as the class with the highest score.
    classIds = []
    confidences = []
    boxes = []
    for out in outs:
        for detection in out:
            scores = detection[5:]
            classId = np.argmax(scores)
            confidence = scores[classId]
            if confidence > confThreshold:
                center_x = int(detection[0] * frameWidth)
                center_y = int(detection[1] * frameHeight)
                width = int(detection[2] * frameWidth)
                height = int(detection[3] * frameHeight)
                left = int(center_x - width / 2)
                top = int(center_y - height / 2)
                classIds.append(classId)
                confidences.append(float(confidence))
                boxes.append([left, top, width, height])

    # Perform non maximum suppression to eliminate redundant overlapping boxes with
    # lower confidences.
    indices = cv2.dnn.NMSBoxes(boxes, confidences, confThreshold, nmsThreshold)
    for i in indices:
        i = i[0]
        box = boxes[i]
        left = box[0]
        top = box[1]
        width = box[2]
        height = box[3]
        # Class "person"
        # rects.append((left, top, left + width, top + height))
        # objects = ct.update(rects)
        # counting(objects)
        # drawPred(classIds[i], confidences[i], left, top, left + width, top + height)
        # person =0
        # bicycle = 1
        # car = 2
        # motorbike = 3
        # bus = 4
        # train = 5
        # truck = 6
        if classIds[i] == 0:
            rects.append((left, top, left + width, top + height))
            # use the centroid tracker to associate the (1) old object
            # centroids with (2) the newly computed object centroids
            objects = ct.update(rects)
            countingObject(objects, trackableObjects, "person", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
        if classIds[i] == 2:
            rects.append((left, top, left + width, top + height))
            # use the centroid tracker to associate the (1) old object
            # centroids with (2) the newly computed object centroids
            objects = ct2.update(rects)
            countingObject(objects, trackableObjects2, "car", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
        if classIds[i] == 3:
            rects.append((left, top, left + width, top + height))
            # use the centroid tracker to associate the (1) old object
            # centroids with (2) the newly computed object centroids
            objects = ct3.update(rects)
            countingObject(objects, trackableObjects3, "motorbike", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
        if classIds[i] == 5:
            rects.append((left, top, left + width, top + height))
            # use the centroid tracker to associate the (1) old object
            # centroids with (2) the newly computed object centroids
            objects = ct4.update(rects)
            countingObject(objects, trackableObjects4, "bus", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)
        if classIds[i] == 6:
            rects.append((left, top, left + width, top + height))
            # use the centroid tracker to associate the (1) old object
            # centroids with (2) the newly computed object centroids
            objects = ct5.update(rects)
            countingObject(objects, trackableObjects5, "truck", frame)

            drawPred(classIds[i], confidences[i], left, top, left + width, top + height, frame)


def countingObject(objects, trackableObjectList, type, frame):
    frameHeight = frame.shape[0]
    frameWidth = frame.shape[1]

    global totalDown
    global totalUp
    global totalUpCar
    global totalUpBus
    global totalUpMotor
    global totalUpTruck
    global totalDownCar
    global totalDownBus
    global totalDownTruck
    global totalDownMotor
    # loop over the tracked objects
    for (objectID, centroid) in objects.items():
        # check to see if a trackable object exists for the current
        # object ID
        to = trackableObjectList.get(objectID, None)

        # if there is no existing trackable object, create one
        if to is None:
            to = TrackableObject(objectID, centroid)

        # otherwise, there is a trackable object so we can utilize it
        # to determine direction
        else:
            # the difference between the y-coordinate of the *current*
            # centroid and the mean of *previous* centroids will tell
            # us in which direction the object is moving (negative for
            # 'up' and positive for 'down')
            y = [c[1] for c in to.centroids]
            direction = centroid[1] - np.mean(y)
            to.centroids.append(centroid)

            # check to see if the object has been counted or not
            if not to.counted:
                # if the direction is negative (indicating the object
                # is moving up) AND the centroid is above the center
                # line, count the object

                if direction < 0 and centroid[1] in range(frameHeight // 2 - 30, frameHeight // 2 + 30):
                    if type is "person":
                        totalUp += 1
                    if type is "car":
                        totalUpCar += 1
                    if type is "bus":
                        totalUpBus += 1
                    if type is "motorbike":
                        totalDownMotor += 1
                    if type is "truck":
                        totalUpTruck += 1
                    to.counted = True

                # if the direction is positive (indicating the object
                # is moving down) AND the centroid is below the
                # center line, count the object
                elif direction > 0 and centroid[1] in range(frameHeight // 2 - 30, frameHeight // 2 + 30):
                    if type is "person":
                        totalDown += 1
                    if type is "car":
                        totalDownCar += 1
                    if type is "bus":
                        totalDownBus += 1
                    if type is "motorbike":
                        totalDownMotor += 1
                    if type is "truck":
                        totalDownTruck += 1
                    to.counted = True

        # store the trackable object in our dictionary
        trackableObjectList[objectID] = to
        # draw both the ID of the object and the centroid of the
        # object on the output frame
        # text = "ID {}".format(objectID)
        # cv.putText(frame, text, (centroid[0] - 10, centroid[1] - 10),
        # cv.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
        cv2.circle(frame, (centroid[0], centroid[1]), 4, (0, 255, 0), -1)
    # construct a tuple of information we will be displaying on the
    # frame
    info = [
        ("Person :", str(totalUp) + "/" + str(totalDown)),
        ("Car  :", str(totalUpCar) + "/" + str(totalDownCar)),
        ("Bus  :", str(totalUpBus) + "/" + str(totalDownBus)),
        ("Motorbike  :", str(totalUpMotor) + "/" + str(totalDownMotor)),
        ("Truck :", str(totalUpTruck) + "/" + str(totalDownTruck)),
    ]

    # loop over the info tuples and draw them on our frame
    for (i, (k, v)) in enumerate(info):
        text = "{}: {}".format(k, v)
        cv2.putText(frame, text, (10, frameHeight - ((i * 60) + 60)),
                    cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 255, 255), 5)


def search_video():
    video = filedialog.askopenfilename()
    if video:
        global cap
        global fps
        global frame_count
        global duration
        cap = cv2.VideoCapture(video)
        fps = cap.get(cv2.CAP_PROP_FPS)  # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        duration = frame_count / fps
        detect_objectsYolo()


def detect_objectsYolo():
    print("ato")
    if cap.isOpened():
        hasFrame, frame = cap.read()
        frame = cv2.resize(frame, (800, 600))
        frameHeight = frame.shape[0]
        frameWidth = frame.shape[1]
        cv2.line(frame, (0, frameHeight // 2), (frameWidth, frameHeight // 2), (0, 255, 255), 2)

        # Create a 4D blob from a frame.
        blob = cv2.dnn.blobFromImage(frame, 1 / 255, (inpWidth, inpHeight), [0, 0, 0], 1, crop=False)

        # Sets the input to the network
        net.setInput(blob)

        # Runs the forward pass to get output of the output layers
        outs = net.forward(getOutputsNames(net))

        # Remove the bounding boxes with low confidence
        postprocess2(frame, outs)

        # Put efficiency information. The function getPerfProfile returns the overall time for inference(t) and the timings for each of the layers(in layersTimes)
        t, _ = net.getPerfProfile()
        label = 'Inference time: %.2f ms' % (t * 1000.0 / cv2.getTickFrequency())
        cv2.putText(frame, label, (0, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

        # show the output image
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        tkimage = ImageTk.PhotoImage(img)
        display1.imgtk = tkimage  # Shows frame for display 1
        display1.configure(image=tkimage)
        # boucle()
        # threadDetection()
        display1.update()
        window.after(10, detect_objectsYolo)
    else:
        ExportExcel()

def threadN():
    # Call work function
    t1 = Thread(target=threadExcel())
    t1.start()
def threadExcel():
    # Call work function
    t1 = Thread(target=ExportExcel())
    t1.start()

def threadSearch():
    t2 = Thread(target=search_video())
    t2.start()


def boucle():
    te = Thread(target=detect_objectsYolo())
    te.start()


def ExportExcel():
    print("Click resultats comptage")
    workbook = Workbook()
    sheet = workbook.active
    totalEntree = 0
    totalSortie = 0
    sheet["A1"] = "LIEU"
    sheet["B1"] = "VIDEO"
    sheet["C1"] = "CAR"
    sheet["D1"] = "BUS"
    sheet["E1"] = "TRUCK"
    sheet["F1"] = "DUREE CAPTURE"
    sheet["G1"] = "ENTREE"
    sheet["H1"] = "1H"
    sheet["I1"] = "14H"
    sheet["J1"] = "SORTIE"
    sheet["K1"] = "1H"
    sheet["L1"] = "14H"
    sheet["M1"] = "TOTAL TRAFFIC"

    sheet["A2"] = "LIEU"
    sheet["B2"] = video
    sheet["C2"] = totalUpCar + totalDownCar
    sheet["D2"] = totalUpBus + totalDownBus
    sheet["E2"] = totalUpTruck + totalDownTruck
    sheet["F2"] = int(duration / 60)
    sheet["G2"] = totalDownCar + totalDownTruck + totalDownBus
    if totalDownCar != 0 or totalDownTruck != 0 or totalDownBus != 0:
        totalEntree = int(((totalDownCar + totalDownTruck + totalDownBus) * 840) / int(duration / 60))
        sheet["H2"] = int(((totalDownCar + totalDownTruck + totalDownBus) * 60) / int(duration / 60))
        sheet["I2"] = totalEntree

    else:
        sheet["H2"] = 0
        sheet["I2"] = 0
    sheet["J2"] = totalUpCar + totalUpTruck + totalUpBus
    if totalUpCar != 0 or totalUpTruck != 0 or totalUpBus != 0:
        totalSortie = int(((totalUpCar + totalUpTruck + totalUpBus) * 840) / int(duration / 60))
        sheet["K2"] = int(((totalUpCar + totalUpTruck + totalUpBus) * 60) / int(duration / 60))
        sheet["L2"] = totalSortie

    else:
        sheet["K2"] = 0
        sheet["L2"] = 0
    ##FFFF00 jaune
    redFill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00',
                          fill_type='solid')
    sheet["M2"].fill = redFill
    sheet["M2"] = totalSortie + totalEntree

    workbook.save(filename="ResultatsDuComptage.xlsx")


def createNewWindow():
    print("Nclick")
    newWindow = tk.Toplevel(window)

    labelExample = tk.Label(newWindow, text="New Window")
    buttonExample = Button(newWindow, text="Exporter Excel", command=threadExcel)

    labelExample.pack()
    buttonExample.pack()


# Set up GUI
window = tk.Tk()  # Makes main window
window.wm_title("Comptage de vehicule")
window.config(background="#FFFFFF")
window.resizable(True, True)
window.geometry("600x500")
# Graphics window
imageFrame = tk.Frame(window, width=600, height=500)
imageFrame.grid(row=0, column=0, padx=10, pady=2)

menubar = Menu(window)
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="Import Video", command=threadSearch)
filemenu.add_command(label="Export", command=threadN)
menubar.add_cascade(label="File", menu=filemenu)
window.config(menu=menubar)


# Capture video frames

# cap = cv2.VideoCapture("video.mp4")


def show_frame():
    _, frame = cap.read()
    frame = cv2.flip(frame, 1)
    cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
    img = Image.fromarray(cv2image)
    imgtk = ImageTk.PhotoImage(image=img)
    display1.imgtk = imgtk  # Shows frame for display 1
    display1.configure(image=imgtk)
    window.after(10, show_frame)


display1 = tk.Label(imageFrame)
display1.grid(row=1, column=0, padx=10, pady=2)
# button = Button(window,                text='Submit',font="Arial 36",                command=search_video())
# button.pack()
# Slider window (slider controls stage position)
# sliderFrame = tk.Frame(window, width=600, height=100)
# sliderFrame.grid(row = 600, column=0, padx=10, pady=2)
#button = Button(window, text="CHOOSE", font="Arial 36", command=threadSearch)
#button2 = Button(window, text="CHOOSE", font="Arial 36", command=threadN)

# detect_objectsYolo()
# show_frame() #Display
window.mainloop()  # Starts GUI
